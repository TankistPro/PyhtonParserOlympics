import time

from typing import Union
from openpyxl import Workbook
from selenium import webdriver
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side
from bs4 import Tag, ResultSet, BeautifulSoup as bs

class CustomOlympicParser:
    """Custom parser for parsing olympics"""

    def getRequestPage(self, url : str) -> Union[str, None]:
        """Getting HTML-markup from url. If URL bad - returned None"""

        service = webdriver.ChromeService()
        driver = webdriver.Chrome(service=service)
        driver.get(url)

        for i in range(20):
            driver.execute_script(
                "window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)

        page = driver.page_source

        if page != "":
            return page


    def getNodeElementList(self, html_markup: str) -> Union[ResultSet, None]:
        """Getting node list from soup"""

        if html_markup:
            soup = bs(markup=html_markup, features="html.parser")
            return soup.find_all("div", class_="fav_olimp olimpiada")

    def parseDataFromNodeList(self, nodeList: dict) -> list:
        """Prepares data to save from node list"""

        response = []

        for node in nodeList:
            title = node.find("span", class_="headline")
            raiting = node.find("span", class_="pl_rating")
            description = node.find("a", class_="none_a black olimp_desc")
            studentClasses = node.find("span", class_="classes_dop")
            subjectTags = node.find_all("span", class_="subject_tag")
            url = node.find("a", class_="none_a black").get_attribute_list("href")[0]
            
            olimpicData = {
                "Название олимпиады": title.text if title is not None else "",
                "Рейтинг": raiting.text if raiting is not None else "",
                "Описание": description.text if description is not None else "",
                "Классы": studentClasses.text if studentClasses is not None else "",
                "Предмет(-ы)": ", ".join(map(lambda x: x.text.strip(), subjectTags)) if subjectTags is not None else "",
                "Ссылка на олимпиаду": f"https://olimpiada.ru{url}" if url is not None else ""
            }

            response.append(olimpicData)

        return response
    
    def saveDataToExcel(self, fileName: str, sheetName: str, preparedData: list[dict]) -> bool:
        """Saving prepared data to excel"""

        wb = Workbook()
        ws = wb.active

        try:
            ws.title = sheetName
            isCreateTableColumnNames = False
            # Заполнение таблицы данными
            for rowIndex in range(len(preparedData)):
                if not isCreateTableColumnNames:
                        columnNames = list(preparedData[rowIndex].keys())
                        ws.append(columnNames)

                        # Стилизация заголовков таблицы
                        for columnIndex in range(len(columnNames)):
                            cell = ws.cell(row=1, column=columnIndex + 1)
                            cell.fill = PatternFill("solid", fgColor="78c4ff")
                            ws.column_dimensions[get_column_letter(columnIndex + 1)].width = 30

                isCreateTableColumnNames = True
                ws.append(list(preparedData[rowIndex].values()))


            # Стилизация готовой таблицы (выравнивание текста, границы)
            totalColumns = len(preparedData[0])
            totalRows = len(preparedData) + 1
            for rowIndex in range(totalRows):
                for columnIndex in range(totalColumns):
                    currentCell = ws.cell(row=rowIndex + 1,column=columnIndex + 1)
                    currentCell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thins = Side(border_style="medium", color="080808")
                    currentCell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

            wb.save(fileName)
            return True
        
        except Exception as error:

            print(error)
            return False
